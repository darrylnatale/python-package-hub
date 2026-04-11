"""
generate_site.py (v2)

Bug fixes:
- Duplicate errors when category == "general"
- Raw markdown/RST/HTML in description
- requests-specific errors appearing on unrelated web packages

New content:
- clean_description() strips all markup before rendering
- get_dependencies() parses requires_dist
- get_recent_releases() returns last 5 releases with dates
- get_code_example() returns package-specific or category fallback example
- get_runtime_errors() returns package-specific post-install errors

Usage:
    python pipeline/generate_site.py

Run from: systems/python-package-hub/
"""

import html as html_lib
import json
import re
import sys
from datetime import date
from pathlib import Path

from jinja2 import Environment, FileSystemLoader

CACHE_DIR = Path("data/package_cache")
TEMPLATES_DIR = Path("site/templates")
STATIC_DIR = Path("site/static")
OUTPUT_DIR = Path("site/output")
SITE_URL = "https://pythonpackagehub.com"


# ── Import name overrides (pip name → Python import name) ────────────────────

IMPORT_NAME_OVERRIDES = {
    "beautifulsoup4": "bs4",
    "pillow": "PIL",
    "opencv-python": "cv2",
    "scikit-learn": "sklearn",
    "python-dateutil": "dateutil",
    "pyyaml": "yaml",
    "python-dotenv": "dotenv",
    "psycopg2-binary": "psycopg2",
    "azure-identity": "azure.identity",
    "azure-storage-blob": "azure.storage.blob",
    "google-auth": "google.auth",
    "google-api-core": "google.api_core",
    "google-cloud-bigquery": "google.cloud.bigquery",
    "google-cloud-storage": "google.cloud.storage",
    "apache-airflow": "airflow",
    "huggingface-hub": "huggingface_hub",
    "pyqt5": "PyQt5",
    "pyqt6": "PyQt6",
    "pyside6": "PySide6",
    "pypdf2": "PyPDF2",
    "pymupdf": "fitz",
    "py-cpuinfo": "cpuinfo",
    "python-docx": "docx",
    "python-barcode": "barcode",
    "factory-boy": "factory",
    "typing-extensions": "typing_extensions",
    "charset-normalizer": "charset_normalizer",
    "async-timeout": "async_timeout",
    "pydantic-settings": "pydantic_settings",
}


# ── Code examples (package-specific) ─────────────────────────────────────────

PACKAGE_EXAMPLES = {
    "requests": """\
import requests

r = requests.get("https://api.github.com/repos/psf/requests")
print(r.status_code)      # 200
data = r.json()
print(data["full_name"])  # psf/requests""",

    "numpy": """\
import numpy as np

arr = np.array([1, 2, 3, 4, 5])
print(arr.mean())   # 3.0
print(arr ** 2)     # [ 1  4  9 16 25]

matrix = np.zeros((3, 3))
print(matrix.shape) # (3, 3)""",

    "pandas": """\
import pandas as pd

df = pd.DataFrame({"name": ["Alice", "Bob"], "age": [30, 25]})
print(df)
print(df["age"].mean())  # 27.5

# Read from CSV
# df = pd.read_csv("data.csv")""",

    "flask": """\
from flask import Flask

app = Flask(__name__)

@app.route("/")
def index():
    return "Hello, World!"

if __name__ == "__main__":
    app.run(debug=True)  # http://localhost:5000""",

    "fastapi": """\
from fastapi import FastAPI

app = FastAPI()

@app.get("/")
def root():
    return {"message": "Hello, World!"}

# Run with: uvicorn main:app --reload
# Docs at: http://localhost:8000/docs""",

    "sqlalchemy": """\
from sqlalchemy import create_engine, text

engine = create_engine("sqlite:///example.db")

with engine.connect() as conn:
    conn.execute(text("CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY, name TEXT)"))
    conn.execute(text("INSERT INTO users (name) VALUES ('Alice')"))
    conn.commit()
    rows = conn.execute(text("SELECT * FROM users")).fetchall()
    print(rows)  # [(1, 'Alice')]""",

    "pydantic": """\
from pydantic import BaseModel

class User(BaseModel):
    id: int
    name: str
    email: str

user = User(id=1, name="Alice", email="alice@example.com")
print(user.model_dump())""",

    "click": """\
import click

@click.command()
@click.option("--name", default="World", help="Who to greet")
def hello(name):
    click.echo(f"Hello, {name}!")

if __name__ == "__main__":
    hello()
# Usage: python script.py --name Alice""",

    "rich": """\
from rich.console import Console
from rich.table import Table

console = Console()
console.print("[bold green]Hello, World![/bold green]")

table = Table(title="Users")
table.add_column("Name")
table.add_column("Age")
table.add_row("Alice", "30")
console.print(table)""",

    "loguru": """\
from loguru import logger

logger.info("Starting application")
logger.warning("Something looks odd")
logger.error("Something went wrong")

# Log to a rotating file
logger.add("app.log", rotation="10 MB")
logger.info("This also writes to app.log")""",

    "celery": """\
from celery import Celery

app = Celery("tasks", broker="redis://localhost:6379/0")

@app.task
def add(x, y):
    return x + y

# Call the task asynchronously
result = add.delay(4, 6)
print(result.get())  # 10
# Start worker: celery -A tasks worker --loglevel=info""",

    "redis": """\
import redis

r = redis.Redis(host="localhost", port=6379, db=0)

r.set("key", "value")
print(r.get("key"))      # b'value'

r.incr("counter")
print(r.get("counter"))  # b'1'""",

    "pillow": """\
from PIL import Image

img = Image.open("photo.jpg")
print(img.size)   # (width, height)
print(img.mode)   # RGB

thumb = img.resize((200, 200))
thumb.save("thumbnail.jpg")""",

    "matplotlib": """\
import matplotlib.pyplot as plt
import numpy as np

x = np.linspace(0, 2 * np.pi, 100)
plt.plot(x, np.sin(x), label="sin(x)")
plt.plot(x, np.cos(x), label="cos(x)")
plt.legend()
plt.title("Trigonometric Functions")
plt.show()""",

    "scikit-learn": """\
from sklearn.datasets import load_iris
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split

iris = load_iris()
X_train, X_test, y_train, y_test = train_test_split(
    iris.data, iris.target, test_size=0.2
)
clf = RandomForestClassifier()
clf.fit(X_train, y_train)
print(clf.score(X_test, y_test))  # ~0.97""",

    "scipy": """\
from scipy import stats
import numpy as np

data = np.array([2.1, 2.5, 2.9, 3.1, 3.5, 3.9, 4.2])
t_stat, p_value = stats.ttest_1samp(data, popmean=3.0)
print(f"t={t_stat:.3f}, p={p_value:.3f}")""",

    "beautifulsoup4": """\
import requests
from bs4 import BeautifulSoup

html = requests.get("https://example.com").text
soup = BeautifulSoup(html, "html.parser")

title = soup.find("title").text
links = [a["href"] for a in soup.find_all("a", href=True)]
print(title, links)""",

    "httpx": """\
import httpx

# Synchronous
r = httpx.get("https://api.github.com")
print(r.status_code)

# Async
import asyncio

async def main():
    async with httpx.AsyncClient() as client:
        r = await client.get("https://api.github.com")
        print(r.json())

asyncio.run(main())""",

    "aiohttp": """\
import aiohttp
import asyncio

async def fetch():
    async with aiohttp.ClientSession() as session:
        async with session.get("https://api.github.com") as r:
            data = await r.json()
            print(data["current_user_url"])

asyncio.run(fetch())""",

    "pytest": """\
# test_math.py
def add(x, y):
    return x + y

def test_add():
    assert add(2, 3) == 5
    assert add(-1, 1) == 0

def test_add_strings():
    assert add("a", "b") == "ab"

# Run: pytest test_math.py -v""",

    "tqdm": """\
from tqdm import tqdm
import time

for i in tqdm(range(100), desc="Processing"):
    time.sleep(0.01)

# With a list comprehension
items = list(range(1000))
results = [x ** 2 for x in tqdm(items)]""",

    "boto3": """\
import boto3

# List S3 buckets
s3 = boto3.client("s3")
response = s3.list_buckets()
for bucket in response["Buckets"]:
    print(bucket["Name"])

# Upload a file
s3.upload_file("local.txt", "my-bucket", "remote.txt")""",

    "psycopg2-binary": """\
import psycopg2

conn = psycopg2.connect(
    host="localhost", database="mydb",
    user="postgres", password="secret"
)
cur = conn.cursor()
cur.execute("SELECT version();")
print(cur.fetchone())
conn.close()""",

    "pymongo": """\
from pymongo import MongoClient

client = MongoClient("mongodb://localhost:27017/")
db = client["mydb"]
col = db["users"]

col.insert_one({"name": "Alice", "age": 30})
for doc in col.find():
    print(doc)""",

    "apscheduler": """\
from apscheduler.schedulers.blocking import BlockingScheduler

scheduler = BlockingScheduler()

@scheduler.scheduled_job("interval", seconds=10)
def job():
    print("Running every 10 seconds")

scheduler.start()""",

    "python-dotenv": """\
from dotenv import load_dotenv
import os

load_dotenv()  # reads .env file in current directory

api_key = os.environ.get("API_KEY")
print(api_key)

# .env file contents:
# API_KEY=your-secret-key
# DEBUG=true""",

    "pyyaml": """\
import yaml

data = {"name": "Alice", "age": 30, "active": True}

# Serialize to YAML string
text = yaml.dump(data)
print(text)

# Parse YAML back to dict
parsed = yaml.safe_load(text)
print(parsed["name"])  # Alice""",

    "typer": """\
import typer

app = typer.Typer()

@app.command()
def greet(name: str, loud: bool = False):
    msg = f"Hello {name}"
    typer.echo(msg.upper() if loud else msg)

if __name__ == "__main__":
    app()
# Usage: python main.py Alice --loud""",

    "faker": """\
from faker import Faker

fake = Faker()
print(fake.name())          # Random full name
print(fake.email())         # Random email
print(fake.address())       # Random address

# Localized data
fake_de = Faker("de_DE")
print(fake_de.city())""",

    "polars": """\
import polars as pl

df = pl.DataFrame({
    "name": ["Alice", "Bob", "Charlie"],
    "age": [30, 25, 35],
    "score": [88.5, 92.0, 78.3],
})
print(df.filter(pl.col("age") > 28))""",

    "anthropic": """\
import anthropic

client = anthropic.Anthropic()  # uses ANTHROPIC_API_KEY env var

message = client.messages.create(
    model="claude-opus-4-6",
    max_tokens=1024,
    messages=[{"role": "user", "content": "Hello, Claude!"}]
)
print(message.content[0].text)""",

    "openai": """\
from openai import OpenAI

client = OpenAI()  # uses OPENAI_API_KEY env var

response = client.chat.completions.create(
    model="gpt-4o",
    messages=[{"role": "user", "content": "Hello!"}]
)
print(response.choices[0].message.content)""",

    "transformers": """\
from transformers import pipeline

# Sentiment analysis
classifier = pipeline("sentiment-analysis")
result = classifier("I love this library!")
print(result)  # [{'label': 'POSITIVE', 'score': 0.9998}]""",

    "torch": """\
import torch

device = "cuda" if torch.cuda.is_available() else "cpu"
print(f"Using: {device}")

x = torch.tensor([1.0, 2.0, 3.0]).to(device)
print(x.mean())   # tensor(2.)""",

    "openpyxl": """\
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Sales"
ws["A1"] = "Month"
ws["B1"] = "Revenue"
ws.append(["January", 12500])
ws.append(["February", 14200])
wb.save("report.xlsx")""",

    "psutil": """\
import psutil

print(f"CPU: {psutil.cpu_percent()}%")
mem = psutil.virtual_memory()
print(f"RAM: {mem.used / 1e9:.1f}GB / {mem.total / 1e9:.1f}GB")
disk = psutil.disk_usage("/")
print(f"Disk free: {disk.free / 1e9:.0f}GB")""",

    "networkx": """\
import networkx as nx

G = nx.Graph()
G.add_edges_from([(1, 2), (1, 3), (2, 4), (3, 4), (4, 5)])
print(f"Nodes: {G.number_of_nodes()}")  # 5
print(f"Edges: {G.number_of_edges()}")  # 5
print(nx.shortest_path(G, 1, 5))       # [1, 2, 4, 5]""",

    "sympy": """\
from sympy import symbols, solve, diff

x = symbols("x")
expr = x**2 - 5*x + 6

print(solve(expr, x))   # [2, 3]
print(diff(expr, x))    # 2*x - 5""",

    "asyncpg": """\
import asyncio
import asyncpg

async def main():
    conn = await asyncpg.connect("postgresql://user:pass@localhost/mydb")
    rows = await conn.fetch("SELECT id, name FROM users LIMIT 5")
    for row in rows:
        print(dict(row))
    await conn.close()

asyncio.run(main())""",

    "aiofiles": """\
import aiofiles
import asyncio

async def main():
    async with aiofiles.open("data.txt", "w") as f:
        await f.write("Hello, async world!")
    async with aiofiles.open("data.txt") as f:
        print(await f.read())

asyncio.run(main())""",

    "streamlit": """\
import streamlit as st
import pandas as pd

st.title("My Data App")
file = st.file_uploader("Upload CSV")
if file:
    df = pd.read_csv(file)
    st.dataframe(df)
    st.line_chart(df.select_dtypes("number"))
# Run: streamlit run app.py""",

    "humanize": """\
import humanize
from datetime import datetime, timedelta

print(humanize.intcomma(1234567))     # 1,234,567
print(humanize.naturalsize(1048576))  # 1.0 MB
past = datetime.now() - timedelta(minutes=5)
print(humanize.naturaltime(past))     # 5 minutes ago""",

    "tabulate": """\
from tabulate import tabulate

data = [["Alice", 30, "Engineer"], ["Bob", 25, "Designer"]]
headers = ["Name", "Age", "Role"]
print(tabulate(data, headers=headers, tablefmt="grid"))""",

    "invoke": """\
from invoke import task

@task
def build(c):
    c.run("python -m build")

@task
def test(c):
    c.run("pytest tests/")

@task(pre=[build])
def deploy(c):
    c.run("scp dist/*.whl server:/opt/app/")
# Run: invoke test""",
}


# ── Category fallback examples ────────────────────────────────────────────────

CATEGORY_EXAMPLES = {
    "web": """\
# Example: make an HTTP request
import requests

response = requests.get("https://httpbin.org/get")
print(response.status_code)
print(response.json())""",

    "data": """\
# Example: create and inspect a DataFrame
import pandas as pd

df = pd.DataFrame({"x": [1, 2, 3], "y": [4, 5, 6]})
print(df.describe())""",

    "ml": """\
# Example: train a simple classifier
from sklearn.linear_model import LogisticRegression
import numpy as np

X = np.array([[0, 0], [1, 1], [1, 0], [0, 1]])
y = [0, 1, 1, 0]
clf = LogisticRegression().fit(X, y)
print(clf.predict([[0.5, 0.5]]))""",

    "testing": """\
# Example: write a pytest test
def multiply(a, b):
    return a * b

def test_multiply():
    assert multiply(3, 4) == 12
    assert multiply(0, 99) == 0
# Run: pytest""",

    "cli": """\
# Example: create a CLI tool with click
import click

@click.command()
@click.argument("name")
def greet(name):
    click.echo(f"Hello, {name}!")

if __name__ == "__main__":
    greet()""",

    "async": """\
# Example: async HTTP fetch
import asyncio
import aiohttp

async def main():
    async with aiohttp.ClientSession() as session:
        async with session.get("https://httpbin.org/get") as r:
            print(await r.json())

asyncio.run(main())""",

    "general": """\
import {import_name}

# Check the installed version
print({import_name}.__version__)""",
}


# ── Package-specific runtime errors ──────────────────────────────────────────

PACKAGE_RUNTIME_ERRORS = {
    "requests": [
        {
            "error": "requests.exceptions.ConnectionError",
            "cause": "The server is unreachable, the URL is wrong, or the network is down.",
            "fix": "Verify the URL is correct and the server is up. Check <code>HTTP_PROXY</code> / <code>HTTPS_PROXY</code> if behind a proxy.",
        },
        {
            "error": "requests.exceptions.Timeout",
            "cause": "The server did not respond within the timeout window.",
            "fix": "Set an explicit timeout: <code>requests.get(url, timeout=10)</code>",
        },
        {
            "error": "requests.exceptions.HTTPError (4xx / 5xx)",
            "cause": "The server returned an error status code.",
            "fix": "Call <code>response.raise_for_status()</code> to raise on error, or check <code>response.status_code</code> first.",
        },
        {
            "error": "requests.exceptions.TooManyRedirects",
            "cause": "The URL redirected more than 30 times.",
            "fix": "Check for redirect loops. Debug with <code>allow_redirects=False</code>.",
        },
    ],
    "pandas": [
        {
            "error": "KeyError: 'column_name'",
            "cause": "The column does not exist in the DataFrame.",
            "fix": "Inspect columns with <code>df.columns.tolist()</code>. Watch for leading/trailing whitespace in column names.",
        },
        {
            "error": "ValueError: Cannot convert float NaN to integer",
            "cause": "The column contains NaN values but is being cast to int.",
            "fix": "Fill or drop NaN values first: <code>df['col'].fillna(0).astype(int)</code>",
        },
        {
            "error": "MemoryError",
            "cause": "The dataset is too large to fit in RAM.",
            "fix": "Use chunked reading: <code>pd.read_csv(file, chunksize=10000)</code> or switch to Polars or Dask.",
        },
        {
            "error": "pandas.errors.ParserError: Error tokenizing data",
            "cause": "Inconsistent column counts or wrong delimiter in the CSV.",
            "fix": "Try <code>pd.read_csv(file, on_bad_lines='skip')</code> or specify <code>sep=','</code> explicitly.",
        },
    ],
    "numpy": [
        {
            "error": "ValueError: operands could not be broadcast together with shapes",
            "cause": "Two arrays have incompatible shapes for the operation.",
            "fix": "Check shapes with <code>a.shape</code> and <code>b.shape</code>. Use <code>.reshape()</code> or <code>np.expand_dims()</code> to align them.",
        },
        {
            "error": "TypeError: ufunc did not contain a loop with signature matching types",
            "cause": "Incompatible data types between arrays (e.g., mixing int and complex).",
            "fix": "Cast to a compatible type: <code>arr.astype(np.float64)</code>",
        },
    ],
    "flask": [
        {
            "error": "werkzeug.exceptions.NotFound (404)",
            "cause": "The requested route is not registered.",
            "fix": "Check route definitions with <code>app.url_map</code>. Verify the HTTP method (GET vs POST) matches.",
        },
        {
            "error": "RuntimeError: Working outside of application context",
            "cause": "Flask globals (g, request, current_app) accessed outside a request context.",
            "fix": "Wrap code with <code>with app.app_context():</code>",
        },
        {
            "error": "OSError: [Errno 98] Address already in use",
            "cause": "Port 5000 is occupied by another process.",
            "fix": "Use a different port: <code>app.run(port=5001)</code>, or find and kill the occupying process.",
        },
    ],
    "fastapi": [
        {
            "error": "HTTP 422 Unprocessable Entity",
            "cause": "Request body or query parameters failed Pydantic validation.",
            "fix": "Check the response body — the <code>detail</code> field lists which fields failed and why.",
        },
        {
            "error": "pydantic_core.ValidationError",
            "cause": "Input data does not match the model's type annotations.",
            "fix": "Inspect <code>e.errors()</code> for field-level details.",
        },
        {
            "error": "ModuleNotFoundError: No module named 'uvicorn'",
            "cause": "uvicorn is a separate package and is not installed.",
            "fix": "Install with: <code>pip install fastapi[standard]</code> (bundles uvicorn) or <code>pip install uvicorn</code>",
        },
    ],
    "sqlalchemy": [
        {
            "error": "sqlalchemy.exc.OperationalError: no such table",
            "cause": "The table has not been created in the database.",
            "fix": "Run <code>Base.metadata.create_all(engine)</code> before using any ORM models.",
        },
        {
            "error": "sqlalchemy.exc.IntegrityError: UNIQUE constraint failed",
            "cause": "Inserting a record violates a unique constraint.",
            "fix": "Check for existing records before insert, or handle the exception and update instead.",
        },
        {
            "error": "sqlalchemy.orm.exc.NoResultFound",
            "cause": "<code>.one()</code> returned zero rows.",
            "fix": "Use <code>.one_or_none()</code> to return None instead of raising, or <code>.first()</code>.",
        },
    ],
    "pytest": [
        {
            "error": "fixture 'X' not found",
            "cause": "Fixture not in scope, misspelled, or the plugin providing it is not installed.",
            "fix": "Define the fixture in <code>conftest.py</code>. For plugin fixtures (e.g., <code>pytest-asyncio</code>), install the plugin.",
        },
        {
            "error": "ImportError while collecting test file",
            "cause": "pytest cannot import the test module.",
            "fix": "Run <code>python -c 'import your_module'</code> to reproduce the error directly.",
        },
    ],
    "click": [
        {
            "error": "click.exceptions.UsageError: Missing argument",
            "cause": "A required argument was not provided on the command line.",
            "fix": "Pass the argument or add <code>required=False</code> with a default value.",
        },
        {
            "error": "click.exceptions.BadParameter",
            "cause": "A parameter value failed type conversion or custom validation.",
            "fix": "Check the expected type. Use <code>type=click.INT</code>, <code>click.FLOAT</code>, etc. for automatic conversion.",
        },
    ],
    "boto3": [
        {
            "error": "botocore.exceptions.NoCredentialsError",
            "cause": "AWS credentials are not configured.",
            "fix": "Run <code>aws configure</code>, set <code>AWS_ACCESS_KEY_ID</code> / <code>AWS_SECRET_ACCESS_KEY</code> env vars, or attach an IAM role.",
        },
        {
            "error": "botocore.exceptions.ClientError: AccessDenied",
            "cause": "The IAM user or role lacks permission for the operation.",
            "fix": "Add the required IAM permission (e.g., <code>s3:GetObject</code>) to the user or role policy.",
        },
        {
            "error": "botocore.exceptions.EndpointConnectionError",
            "cause": "Wrong region specified or no network access to the AWS endpoint.",
            "fix": "Set the correct region: <code>boto3.client('s3', region_name='us-east-1')</code>",
        },
    ],
    "redis": [
        {
            "error": "redis.exceptions.ConnectionError: Error connecting to localhost:6379",
            "cause": "Redis server is not running or not reachable.",
            "fix": "Start Redis: <code>redis-server</code> or <code>docker run -p 6379:6379 redis</code>",
        },
        {
            "error": "redis.exceptions.ResponseError: WRONGTYPE Operation against a key",
            "cause": "The command is incompatible with the data type stored at that key.",
            "fix": "Use the correct command for the stored type, or delete the key with <code>r.delete('key')</code>.",
        },
    ],
    "celery": [
        {
            "error": "kombu.exceptions.OperationalError: Connection refused",
            "cause": "The message broker (Redis or RabbitMQ) is not running.",
            "fix": "Start the broker: <code>redis-server</code>. Check the broker URL in your Celery config.",
        },
        {
            "error": "celery.exceptions.NotRegistered: task_name",
            "cause": "The task module was not imported before the worker started.",
            "fix": "Use <code>app.autodiscover_tasks(['myapp'])</code> or import task modules explicitly in <code>celery.py</code>.",
        },
    ],
    "pydantic": [
        {
            "error": "pydantic.ValidationError",
            "cause": "Input data does not match the model's type annotations or validators.",
            "fix": "Call <code>e.errors()</code> for field-level detail. Each entry shows location, type, and message.",
        },
        {
            "error": "TypeError: unexpected keyword argument",
            "cause": "A field name passed to the model constructor doesn't exist in the model definition.",
            "fix": "Check model field names match what you're passing. Use <code>model_config = ConfigDict(extra='forbid')</code> to catch this at definition time.",
        },
    ],
    "pillow": [
        {
            "error": "OSError: cannot identify image file",
            "cause": "The file is not a valid image, is corrupted, or the path is wrong.",
            "fix": "Verify the path exists with <code>os.path.exists(path)</code>. Confirm it's a valid image by opening it in a viewer.",
        },
        {
            "error": "PIL.UnidentifiedImageError",
            "cause": "Pillow cannot determine the image format from the file header.",
            "fix": "Ensure the file is not truncated or zero bytes. Check if the extension matches the actual format.",
        },
    ],
    "selenium": [
        {
            "error": "WebDriverException: 'chromedriver' executable needs to be in PATH",
            "cause": "ChromeDriver binary is not installed or not on PATH.",
            "fix": "Install automatically: <code>pip install webdriver-manager</code> and use <code>ChromeDriverManager().install()</code>",
        },
        {
            "error": "NoSuchElementException",
            "cause": "The element was not found on the page — it may not have loaded yet.",
            "fix": "Use explicit waits: <code>WebDriverWait(driver, 10).until(EC.presence_of_element_located(...))</code>",
        },
        {
            "error": "StaleElementReferenceException",
            "cause": "The DOM was updated after the element reference was captured.",
            "fix": "Re-find the element after any page action that triggers a DOM update.",
        },
    ],
    "aiohttp": [
        {
            "error": "aiohttp.ClientConnectorError: Cannot connect to host",
            "cause": "Server unreachable or URL incorrect.",
            "fix": "Verify the URL and check network connectivity. Ensure the server is running if testing locally.",
        },
        {
            "error": "RuntimeError: Session is closed",
            "cause": "The ClientSession was used after being closed.",
            "fix": "Use <code>async with aiohttp.ClientSession() as session:</code> to manage the session lifecycle.",
        },
    ],
    "apscheduler": [
        {
            "error": "apscheduler.jobstores.base.JobLookupError",
            "cause": "Attempting to modify or remove a job ID that does not exist.",
            "fix": "List current jobs with <code>scheduler.get_jobs()</code>. Check the job ID is correct.",
        },
        {
            "error": "apscheduler.jobstores.base.ConflictingIdError",
            "cause": "A job with that ID is already registered.",
            "fix": "Use a unique job ID, or pass <code>replace_existing=True</code> when adding the job.",
        },
    ],
    "psycopg2-binary": [
        {
            "error": "psycopg2.OperationalError: could not connect to server: Connection refused",
            "cause": "PostgreSQL is not running or not accepting connections.",
            "fix": "Start PostgreSQL and verify host, port, database name, user, and password.",
        },
        {
            "error": "psycopg2.ProgrammingError: relation does not exist",
            "cause": "The table referenced in the SQL query does not exist in this database.",
            "fix": "Run your schema migrations. Confirm you are connected to the right database.",
        },
    ],
    "pymongo": [
        {
            "error": "pymongo.errors.ServerSelectionTimeoutError",
            "cause": "Cannot connect to MongoDB within the timeout period.",
            "fix": "Ensure MongoDB is running (<code>mongod</code>). Default port is 27017.",
        },
        {
            "error": "pymongo.errors.DuplicateKeyError",
            "cause": "Inserting a document violates a unique index constraint.",
            "fix": "Use <code>update_one(..., upsert=True)</code> instead of <code>insert_one</code> to update if it exists.",
        },
    ],
    "openai": [
        {
            "error": "openai.AuthenticationError: Incorrect API key",
            "cause": "The API key is missing, wrong, or revoked.",
            "fix": "Set <code>OPENAI_API_KEY</code> environment variable or pass <code>api_key='sk-...'</code> to the client.",
        },
        {
            "error": "openai.RateLimitError",
            "cause": "Too many requests or usage quota exceeded.",
            "fix": "Add retry logic with backoff. Use the <code>tenacity</code> library or <code>time.sleep()</code> between requests.",
        },
    ],
    "anthropic": [
        {
            "error": "anthropic.AuthenticationError",
            "cause": "The API key is missing or invalid.",
            "fix": "Set the <code>ANTHROPIC_API_KEY</code> environment variable or pass <code>api_key='...'</code> to <code>Anthropic()</code>.",
        },
        {
            "error": "anthropic.RateLimitError",
            "cause": "API rate limit exceeded.",
            "fix": "Add exponential backoff between requests. Reduce request frequency or upgrade to a higher rate limit tier.",
        },
    ],
}


# ── Installation errors ───────────────────────────────────────────────────────
# Uses {pip_name} for pip commands and {import_name} for module names.

INSTALL_ERRORS_BASE = [
    {
        "error": "ModuleNotFoundError: No module named '{import_name}'",
        "cause": "The package is not installed in the current Python environment.",
        "fix": "Run <code>pip install {pip_name}</code>. If using a virtual environment, ensure it is activated first.",
    },
    {
        "error": "ModuleNotFoundError: No module named '{import_name}' (installed but still failing)",
        "cause": "pip installed the package into a different Python than the one running your script.",
        "fix": "Use <code>python -m pip install {pip_name}</code> to install into the interpreter you are running.",
    },
    {
        "error": "ImportError: cannot import name 'X' from '{import_name}'",
        "cause": "The function or class does not exist in the installed version.",
        "fix": "Check the version with <code>pip show {pip_name}</code> and upgrade with <code>pip install --upgrade {pip_name}</code>.",
    },
    {
        "error": "pip: command not found",
        "cause": "pip is not in PATH or Python was not added to PATH during installation.",
        "fix": "Try <code>python -m pip install {pip_name}</code>. On macOS/Linux try <code>pip3</code>.",
    },
    {
        "error": "PermissionError: [Errno 13] Permission denied",
        "cause": "No write access to the system Python package directory.",
        "fix": "Use a virtual environment, or add <code>--user</code>: <code>pip install --user {pip_name}</code>",
    },
    {
        "error": "SSL: CERTIFICATE_VERIFY_FAILED",
        "cause": "pip cannot verify PyPI's SSL certificate — common behind corporate proxies.",
        "fix": "Try: <code>pip install --trusted-host pypi.org --trusted-host files.pythonhosted.org {pip_name}</code>",
    },
]

# Category-specific install errors — intentionally generic (no package-specific API references)
INSTALL_ERRORS_BY_CATEGORY = {
    "web": [
        {
            "error": "ConnectionError: Failed to establish a new connection",
            "cause": "Server unreachable, URL invalid, or firewall/proxy blocking the connection.",
            "fix": "Verify the URL and network access. Set <code>HTTP_PROXY</code> / <code>HTTPS_PROXY</code> env vars if behind a proxy.",
        },
        {
            "error": "SSLError: CERTIFICATE_VERIFY_FAILED",
            "cause": "The remote server's SSL certificate cannot be verified.",
            "fix": "Update CA certificates on your system. For testing only, disable SSL verification (never in production).",
        },
    ],
    "data": [
        {
            "error": "MemoryError when loading data",
            "cause": "Dataset is too large to fit in RAM.",
            "fix": "Read in chunks, filter columns on load, or consider Polars/Dask for out-of-core processing.",
        },
    ],
    "ml": [
        {
            "error": "RuntimeError: CUDA out of memory",
            "cause": "Model or batch size exceeds GPU memory.",
            "fix": "Reduce batch size, call <code>torch.cuda.empty_cache()</code>, or run on CPU.",
        },
    ],
    "testing": [],
    "cli": [],
    "async": [
        {
            "error": "RuntimeError: This event loop is already running",
            "cause": "Calling <code>asyncio.run()</code> inside Jupyter or another already-running event loop.",
            "fix": "In Jupyter, use <code>await</code> directly. Or install <code>nest_asyncio</code> and call <code>nest_asyncio.apply()</code>.",
        },
    ],
    "general": [],
}


# ── Category detection ────────────────────────────────────────────────────────

CATEGORY_CLASSIFIERS = {
    "web": ["Topic :: Internet :: WWW/HTTP", "Framework :: Django", "Framework :: Flask", "Framework :: FastAPI"],
    "data": ["Topic :: Scientific/Engineering", "Topic :: Scientific/Engineering :: Information Analysis"],
    "ml": ["Topic :: Scientific/Engineering :: Artificial Intelligence"],
    "testing": ["Framework :: Pytest", "Topic :: Software Development :: Testing"],
    "cli": ["Environment :: Console", "Topic :: Utilities"],
    "async": ["Framework :: AsyncIO"],
}

CATEGORY_LABELS = {
    "web": "Web & HTTP",
    "data": "Data & Science",
    "ml": "Machine Learning & AI",
    "testing": "Testing & QA",
    "cli": "CLI & Utilities",
    "async": "Async & Networking",
    "general": "General Purpose",
}


def get_category(classifiers, name):
    n = name.lower()
    if any(x in n for x in ["test", "pytest", "mock", "coverage", "hypothesis"]):
        return "testing"
    if any(x in n for x in ["torch", "tensorflow", "keras", "transformers", "sklearn", "xgboost", "lightgbm"]):
        return "ml"
    if any(x in n for x in ["flask", "django", "fastapi", "aiohttp", "starlette", "httpx", "requests", "scrapy"]):
        return "web"
    if any(x in n for x in ["pandas", "numpy", "scipy", "plotly", "matplotlib", "seaborn", "polars", "dask"]):
        return "data"
    if any(x in n for x in ["click", "typer", "rich", "loguru", "colorama", "tabulate"]):
        return "cli"
    if any(x in n for x in ["async", "aio", "websocket", "trio"]):
        return "async"
    for cat, markers in CATEGORY_CLASSIFIERS.items():
        for marker in markers:
            if any(marker in c for c in classifiers):
                return cat
    return "general"


# ── License normalization ─────────────────────────────────────────────────────

# Maps regex patterns (searched in lowercase license text) to SPDX identifiers.
_LICENSE_PATTERNS = [
    (r"\bapache.?2", "Apache-2.0"),
    (r"\bapache.?software.?license", "Apache-2.0"),
    (r"\bgpl.?3", "GPL-3.0"),
    (r"\bgpl.?2", "GPL-2.0"),
    (r"\blgpl.?3", "LGPL-3.0"),
    (r"\blgpl.?2", "LGPL-2.1"),
    (r"\bagpl.?3", "AGPL-3.0"),
    (r"\bmpl.?2", "MPL-2.0"),
    (r"\bbsd.?3", "BSD-3-Clause"),
    (r"\bbsd.?2", "BSD-2-Clause"),
    (r"\bbsd\b", "BSD"),
    (r"\bmit\b", "MIT"),
    (r"\bisc\b", "ISC"),
    (r"\bcc0\b", "CC0-1.0"),
    (r"\bunlicense\b", "Unlicense"),
    (r"\bpython software foundation\b", "PSF"),
    (r"\bpsfl\b", "PSF"),
    (r"\bpsf\b", "PSF"),
    (r"\beupl\b", "EUPL-1.2"),
    (r"\bartistic\b", "Artistic-2.0"),
    (r"\bepl.?2", "EPL-2.0"),
    (r"\bepl.?1", "EPL-1.0"),
    (r"\bcddl\b", "CDDL-1.0"),
]


def normalize_license(raw):
    """Return a short SPDX-style license identifier. Returns '' if unclear."""
    if not raw:
        return ""
    # Already short and on one line — use as-is
    if len(raw) <= 50 and "\n" not in raw:
        return raw.strip()
    # Multi-line or very long — try keyword matching
    raw_lower = raw.lower()
    for pattern, spdx in _LICENSE_PATTERNS:
        if re.search(pattern, raw_lower):
            return spdx
    # Try first non-empty line if it's short and looks like an identifier (not body text)
    first_line = next((l.strip() for l in raw.splitlines() if l.strip()), "")
    if first_line and len(first_line) <= 50 and not first_line.lower().startswith("copyright"):
        return first_line
    # Can't extract a clean identifier — omit rather than show garbage
    return ""


# ── Utility functions ─────────────────────────────────────────────────────────

def get_import_name(pip_name):
    override = IMPORT_NAME_OVERRIDES.get(pip_name.lower())
    if override:
        return override
    return pip_name.lower().replace("-", "_")


def clean_description(text):
    """Strip markdown, RST, and HTML markup. Return list of plain text paragraphs."""
    if not text or len(text.strip()) < 20:
        return []

    # Unescape HTML entities first
    text = html_lib.unescape(text)

    # Strip HTML tags
    text = re.sub(r"<[^>]+>", " ", text)

    # Strip RST directives (.. image:: ...) and field options (:target: ...)
    text = re.sub(r"^\.\.\s+\S.*$", "", text, flags=re.MULTILINE)
    text = re.sub(r"^\s+:\w[\w\s-]*:.*$", "", text, flags=re.MULTILINE)
    # Strip RST section underline lines (====, ----, ~~~~)
    text = re.sub(r"^[=\-~#^*+]{4,}\s*$", "", text, flags=re.MULTILINE)

    # Strip markdown badge images: [![text](url)](url)
    text = re.sub(r"!\[[^\]]*\]\([^\)]*\)(?:\([^\)]*\))?", "", text)
    # Strip markdown links, keep text: [text](url) → text
    text = re.sub(r"\[([^\]]+)\]\([^\)]+\)", r"\1", text)
    # Strip bare reference definitions: [label]: url
    text = re.sub(r"^\[[^\]]+\]:\s+\S.*$", "", text, flags=re.MULTILINE)
    # Strip markdown headings
    text = re.sub(r"^#{1,6}\s+", "", text, flags=re.MULTILINE)
    # Strip markdown bold/italic (allow spanning a single newline)
    text = re.sub(r"\*{1,3}([^*]+?)\*{1,3}", r"\1", text)
    text = re.sub(r"_{1,2}([^_]+?)_{1,2}", r"\1", text)

    # Strip fenced code blocks
    text = re.sub(r"```[\s\S]*?```", "", text)
    text = re.sub(r"~~~[\s\S]*?~~~", "", text)
    # Strip inline code ticks
    text = re.sub(r"`[^`\n]+`", "", text)

    # Strip bare URLs
    text = re.sub(r"https?://\S+", "", text)

    # Strip RST bullet-list items that are just a label followed by a colon with no content.
    # These appear after URL stripping from RST link-reference lists like "- Website: <url>".
    # Pattern: line starting with "- " then a short label ending in ":" with nothing after.
    text = re.sub(r"^-\s+[^\n:]{0,50}:\s*$", "", text, flags=re.MULTILINE)

    # Normalize whitespace
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()

    # License/legal boilerplate detector — skip paragraphs that match 2+ patterns
    license_signals = [
        r"redistribution and use",
        r"permission is hereby granted",
        r"all rights reserved",
        r"without warranty",
        r"the above copyright",
        r"provided by the copyright",
        r"neither the name of",
        r"as-is",
        r"software foundation",
        r"apache license",
        r"mit license",
        r"bsd license",
    ]

    def is_license_para(p):
        p_lower = p.lower()
        return sum(1 for sig in license_signals if sig in p_lower) >= 2

    # Split into paragraphs, filter noise
    paragraphs = []
    for para in re.split(r"\n\n+", text):
        para = para.strip()
        if len(para) < 40 or para.count(" ") < 3:
            continue
        # Skip lines that are mostly special characters (table borders, etc.)
        special_ratio = len(re.findall(r"[|<>{}\[\]\\^~=]", para)) / max(len(para), 1)
        if special_ratio > 0.12:
            continue
        if is_license_para(para):
            continue
        # Skip badge-artifact paragraphs: many newlines relative to word count.
        # These are the remnants of RST/markdown badge tables where image URLs were
        # stripped, leaving isolated words on separate lines (e.g. "docs\n\ntests\n\npackage").
        para_lines = para.count("\n") + 1
        para_words = len(para.split())
        if para_lines > 3 and para_words / para_lines < 1.5:
            continue
        paragraphs.append(para)

    return paragraphs[:3]


def get_install_commands(name):
    return {
        "pip": f"pip install {name}",
        "pip3": f"pip3 install {name}",
        "venv": (
            f"python -m venv venv\n"
            f"source venv/bin/activate   # Windows: venv\\Scripts\\activate\n"
            f"pip install {name}"
        ),
        "conda": f"conda install -c conda-forge {name}",
        "poetry": f"poetry add {name}",
    }


def get_verify_command(import_name):
    return f'python -c "import {import_name}; print({import_name}.__version__)"'


def get_dependencies(data):
    """Return list of core dependency names from requires_dist, filtering extras."""
    requires = data.get("info", {}).get("requires_dist") or []
    deps = []
    for req in requires:
        # Skip anything with extra == condition
        if "extra ==" in req.lower():
            continue
        # Take only the package name (before any version specifier or semicolon)
        name = re.split(r"[;(>=<!~\s]", req)[0].strip()
        if name:
            deps.append(name)
    return deps[:12]


def get_recent_releases(data):
    """Return the 5 most recent releases with dates."""
    releases = data.get("releases", {})
    current = data.get("info", {}).get("version", "")
    result = []
    for version, artifacts in releases.items():
        if not artifacts:
            continue
        dates = [a.get("upload_time", "")[:10] for a in artifacts if a.get("upload_time")]
        if not dates:
            continue
        result.append({
            "version": version,
            "date": min(dates),
            "is_current": version == current,
        })
    result.sort(key=lambda x: x["date"], reverse=True)
    return result[:5]


def get_install_errors(pip_name, import_name, category):
    """Bug fix: skip category extras when category == 'general' (was duplicating errors)."""
    base = INSTALL_ERRORS_BASE
    extra = [] if category == "general" else INSTALL_ERRORS_BY_CATEGORY.get(category, [])
    result = []
    for err in base + extra:
        result.append({
            "error": err["error"].replace("{pip_name}", pip_name).replace("{import_name}", import_name),
            "cause": err["cause"].replace("{pip_name}", pip_name).replace("{import_name}", import_name),
            "fix": err["fix"].replace("{pip_name}", pip_name).replace("{import_name}", import_name),
        })
    return result


def get_runtime_errors(pip_name):
    """Return package-specific runtime errors, or empty list if none defined."""
    return PACKAGE_RUNTIME_ERRORS.get(pip_name.lower(), [])


def get_code_example(pip_name, import_name):
    """Return a package-specific example if one exists, otherwise a minimal
    import-and-version snippet for this specific package. Never borrows examples
    from unrelated packages via category fallback."""
    example = PACKAGE_EXAMPLES.get(pip_name.lower()) or PACKAGE_EXAMPLES.get(pip_name)
    if example:
        return example
    # Generic fallback: minimal import + version check, personalised to this package.
    # Using the import name keeps it accurate (e.g. PIL for pillow, sklearn for scikit-learn).
    return f"import {import_name}\n\nprint({import_name}.__version__)"


# ── Data extraction ───────────────────────────────────────────────────────────

def extract_info(data):
    info = data.get("info", {})
    name = info.get("name", "")
    return {
        "name": name,
        "import_name": get_import_name(name),
        "slug": name.lower().replace(".", "-"),
        "version": info.get("version", ""),
        "summary": (info.get("summary", "") or "").strip(),
        "description_paragraphs": clean_description(info.get("description", "") or ""),
        "author": (info.get("author", "") or info.get("author_email", "") or "").strip(),
        "license": normalize_license((info.get("license", "") or "").strip()),
        "home_page": (info.get("home_page", "") or "").strip(),
        "requires_python": (info.get("requires_python", "") or "").strip(),
        "classifiers": info.get("classifiers", []),
        "project_urls": info.get("project_urls", {}) or {},
    }


# ── Site generation ───────────────────────────────────────────────────────────

def load_package_data(cache_dir):
    packages = []
    for json_file in sorted(cache_dir.glob("*.json")):
        try:
            with open(json_file, encoding="utf-8") as f:
                packages.append(json.load(f))
        except (json.JSONDecodeError, IOError) as e:
            print(f"  [WARN] {json_file.name}: {e}")
    return packages


def generate_pages(packages, env, output_dir):
    template = env.get_template("package.html")
    generated = []

    for data in packages:
        pkg = extract_info(data)
        if not pkg["name"]:
            continue

        category = get_category(pkg["classifiers"], pkg["name"])
        category_label = CATEGORY_LABELS.get(category, "General Purpose")

        page_dir = output_dir / pkg["slug"]
        page_dir.mkdir(parents=True, exist_ok=True)

        html = template.render(
            pkg=pkg,
            category=category,
            category_label=category_label,
            install_commands=get_install_commands(pkg["name"]),
            verify_command=get_verify_command(pkg["import_name"]),
            install_errors=get_install_errors(pkg["name"], pkg["import_name"], category),
            runtime_errors=get_runtime_errors(pkg["name"]),
            code_example=get_code_example(pkg["name"], pkg["import_name"]),
            dependencies=get_dependencies(data),
            recent_releases=get_recent_releases(data),
            site_url=SITE_URL,
            today=date.today().isoformat(),
        )

        (page_dir / "index.html").write_text(html, encoding="utf-8")
        generated.append({
            "name": pkg["name"],
            "slug": pkg["slug"],
            "summary": pkg["summary"],
            "category": category_label,
        })

    return generated


def generate_homepage(packages_meta, env, output_dir):
    template = env.get_template("index.html")
    html = template.render(
        packages=packages_meta,
        total=len(packages_meta),
        site_url=SITE_URL,
        today=date.today().isoformat(),
    )
    (output_dir / "index.html").write_text(html, encoding="utf-8")


def copy_static(static_dir, output_dir):
    static_out = output_dir / "static"
    static_out.mkdir(parents=True, exist_ok=True)
    for f in static_dir.iterdir():
        if f.is_file():
            (static_out / f.name).write_bytes(f.read_bytes())


def main():
    for d in [CACHE_DIR, TEMPLATES_DIR, STATIC_DIR]:
        if not d.exists():
            print(f"ERROR: {d} not found. Run from the system root directory.")
            sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    env = Environment(loader=FileSystemLoader(str(TEMPLATES_DIR)), autoescape=True)

    print("Loading cached package data...")
    packages = load_package_data(CACHE_DIR)
    print(f"  Loaded {len(packages)} packages")

    print("Generating package pages...")
    packages_meta = generate_pages(packages, env, OUTPUT_DIR)
    print(f"  Generated {len(packages_meta)} pages")

    print("Generating homepage...")
    generate_homepage(packages_meta, env, OUTPUT_DIR)

    print("Copying static assets...")
    copy_static(STATIC_DIR, OUTPUT_DIR)

    print(f"\nDone. {len(packages_meta) + 1} pages in {OUTPUT_DIR.resolve()}")


if __name__ == "__main__":
    main()
